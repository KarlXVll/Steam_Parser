
import requests
import openpyxl
from bs4 import BeautifulSoup

# Настройки
steam_api_key = "672BC055EB1F2C8241E1A0012AF6A300"
excel_file_name = "Игры2.xlsx"
excel_file_path = "C:\\Users\\Aboba\\Desktop"

# Функция для поиска информации об игре
def get_game_info(game_name):
    url = f"https://api.steampowered.com/IStoreService/GetAppDetails/?appids={game_name}&key={steam_api_key}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        game_info = {
            "name": data["app"]["name"],
            "publisher": data["app"]["publisher"],
            "genres": ", ".join(data["app"]["genres"]),
            "release_year": data["app"]["release_date"]["year"],
            "price": data["app"]["price"]["initial"] if data["app"]["price"]["initial"] != 0 else "Бесплатно",
            "steam_url": f"https://store.steampowered.com/app/{data['app']['appid']}/"
        }
        return game_info
    else:
        return None

# Функция для записи информации в Excel-таблицу
def write_to_excel(game_info, excel_file_name, excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path + "\\\\" + excel_file_name)
    sheet = wb.active
    
    # Поиск последней строки
    last_row = sheet.max_row
    
    # Запись информации
    sheet.cell(row=last_row + 1, column=1).value = game_info["name"]
    sheet.cell(row=last_row + 1, column=2).value = game_info["publisher"]
    sheet.cell(row=last_row + 1, column=3).value = game_info["genres"]
    sheet.cell(row=last_row + 1, column=4).value = game_info["release_year"]
    sheet.cell(row=last_row + 1, column=5).value = game_info["price"]
    sheet.cell(row=last_row + 1, column=6).value = game_info["steam_url"]
    sheet.cell(row=last_row + 1, column=7).value = "Да" if game_info else "Игра не найдена"
    
    wb.save(excel_file_path + "\\" + excel_file_name)

# Парсинг игр
wb = openpyxl.load_workbook(excel_file_path + "\\" + excel_file_name)
sheet = wb.active

for row in range(2, sheet.max_row + 1):
    game_name = sheet.cell(row=row, column=1).value
    game_info = get_game_info(game_name)
    
    if game_info:
        write_to_excel(game_info, excel_file_name, excel_file_path)
    else:
        print(f"Игра '{game_name}' не найдена в Steam.")

print("Парсинг завершен.")